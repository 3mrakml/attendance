"""
gen_test_excel.py
=================
ينشئ ملف Excel اختباري لخاصية "استيراد الطلاب" في نظام الحضور.

السيناريوهات المغطاة:
  ✅ نجاح  - طالب بكل بياناته (اسم + هاتف + عمر + تاريخ ميلاد + صف موجود)
  ✅ نجاح  - طالب بهاتف مشترك (أخ)
  ✅ نجاح  - طالب بدون هاتف
  ✅ نجاح  - طالب بدون عمر ولا تاريخ ميلاد
  ✅ نجاح  - طالب باسم عربي مع همزات وتاء مربوطة (اختبار تطبيع النص)
  ❌ فشل   - اسم فارغ (خلية الاسم فارغة تماماً)
  ❌ فشل   - صف فارغ (اسم موجود لكن عمود الصف فارغ)
  ❌ فشل   - صف غير موجود في النظام (اسم موجود في الملف بس مش مسجل في النظام)
  ❌ فشل   - تكرار داخل الملف (نفس الصف مكتوب مرتين بنفس البيانات)
  ❌ فشل   - تكرار داخل الملف مع هاتف مختلف (يجب أن ينجح - ليس تكراراً تاماً)
  ❌ فشل   - صف فارغ بالكامل (يُتخطى ولا يُعدّ خطأ)

ملاحظة: السيناريو التاسع (تكرار من قاعدة البيانات) لا يمكن
       محاكاته بدون بيانات موجودة فعلاً في النظام، لذلك
       المصف رقم 11 في الملف مكرر من الصف رقم 2 بنفس البيانات.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "test_students_import.xlsx")

# ─────────────────────────────────────────────────────────────────
# بيانات الصف الموجودة (يجب أن يتطابق الاسم مع ما هو في النظام!)
# غيّر هذه القيم لتتناسب مع الصفوف المسجلة عندك في النظام
# ─────────────────────────────────────────────────────────────────
VALID_GRADE   = "الفرقة الأولى"     # صف موجود في النظام   ✅
INVALID_GRADE = "الفرقة العاشرة"   # صف غير موجود في النظام ❌

# ─────────────────────────────────────────────────────────────────
# البيانات: (الاسم, السن, الهاتف, الصف, تاريخ الميلاد, ملاحظة)
# ─────────────────────────────────────────────────────────────────
ROWS = [
    # ------- طلاب يجب أن ينجحوا ✅ -------
    ("أحمد محمد علي",     20, "01012345678", VALID_GRADE,   "2004-03-15", "✅ طالب مكتمل البيانات"),
    ("منة الله سامح",    19, "01112345678", VALID_GRADE,   "2005-06-20", "✅ طالبة مكتملة"),
    ("خالد إبراهيم",     18, "01012345678", VALID_GRADE,   "2006-01-01", "✅ هاتف مشترك مع الأول (أخوه) - يُقبل"),
    ("سارة عبد الرحمن",  21, "",           VALID_GRADE,   "",           "✅ بدون هاتف ولا تاريخ ميلاد"),
    ("محمود حسن",        22, "01512345678", VALID_GRADE,   "",           "✅ بدون تاريخ ميلاد، عمر يدوي فقط"),
    ("أميرة فؤاد",       20, "01212345678", VALID_GRADE,   "2004-09-05", "✅ اسم بهمزة وتاء مربوطة (اختبار تطبيع)"),

    # ------- طلاب يجب أن يفشلوا ❌ -------

    # الفشل 1: اسم فارغ
    ("",                  18, "01099999999", VALID_GRADE,   "",           "❌ اسم فارغ"),

    # الفشل 2: صف فارغ
    ("يوسف أحمد",         19, "01055555555", "",            "",           "❌ صف فارغ"),

    # الفشل 3: صف غير موجود في النظام
    ("ريم سالم",          20, "01066666666", INVALID_GRADE, "",           "❌ صف غير موجود في النظام"),

    # الفشل 4: تكرار داخل الملف (نفس بيانات الصف الأول حرفياً)
    ("أحمد محمد علي",     20, "01012345678", VALID_GRADE,   "2004-03-15", "❌ تكرار تام للطالب الأول داخل الملف"),

    # الفشل 5: تكرار داخل الملف (نفس البيانات مرة ثالثة)
    ("أحمد محمد علي",     20, "01012345678", VALID_GRADE,   "2004-03-15", "❌ تكرار تام للطالب الأول (مرة ثالثة)"),

    # ملحوظة: الصف الفارغ تماماً يُتخطى تلقائياً (لا يظهر كخطأ)
    # سيتم إدراج صف فارغ في الملف ليُختبر هذا السلوك
]

# ─────────────────────────────────────────────────────────────────
# الألوان
# ─────────────────────────────────────────────────────────────────
COLOR_HEADER  = "3B82F6"  # أزرق (رأس الجدول)
COLOR_SUCCESS = "D1FAE5"  # أخضر فاتح
COLOR_FAIL    = "FEE2E2"  # أحمر فاتح
COLOR_NOTE    = "F8FAFC"  # رمادي جداً فاتح

def make_border():
    thin = Side(style="thin", color="CBD5E1")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def build_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الطلاب"
    ws.sheet_view.rightToLeft = True

    # ── رأس الجدول ────────────────────────────────────────────────
    headers = ["الاسم بالكامل", "السن", "رقم الهاتف", "الصف", "تاريخ الميلاد", "ملاحظة (للاختبار - احذفها قبل الرفع)"]
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True, color="FFFFFF", name="Cairo", size=11)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()

    # ── بيانات الطلاب ────────────────────────────────────────────
    data_rows = []
    for r in ROWS:
        name, age, phone, grade, dob, note = r
        data_rows.append([name, age if age else "", phone, grade, dob, note])

    # صف فارغ بالكامل (السيناريو: يتجاهله النظام بدون خطأ)
    EMPTY_ROW_INDEX = 8  # بعد الطالب السابع
    data_rows.insert(EMPTY_ROW_INDEX - 1, ["", "", "", "", "", "← صف فارغ تماماً (يُتخطى بصمت)"])

    for row_idx, row_data in enumerate(data_rows, start=2):
        note_val = row_data[5] if len(row_data) > 5 else ""
        is_empty = all(str(v).strip() == "" for v in row_data[:5])
        is_success = note_val.startswith("✅")
        is_fail    = note_val.startswith("❌")

        if is_empty:
            row_fill = PatternFill("solid", fgColor="F1F5F9")
        elif is_success:
            row_fill = PatternFill("solid", fgColor=COLOR_SUCCESS)
        elif is_fail:
            row_fill = PatternFill("solid", fgColor=COLOR_FAIL)
        else:
            row_fill = PatternFill("solid", fgColor=COLOR_NOTE)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = make_border()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(name="Cairo", size=10)

            # عمود السن: رقم
            if col_idx == 2 and value != "":
                try:
                    cell.value = int(value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                except (ValueError, TypeError):
                    pass

            # عمود الملاحظة: لون مختلف للنص
            if col_idx == 6:
                cell.font = Font(name="Cairo", size=9, color="64748B", italic=True)

    # ── عرض الأعمدة ───────────────────────────────────────────────
    col_widths = [32, 8, 18, 22, 18, 45]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30

    # ── تجميد الصف الأول ──────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── ورقة ثانية: السيناريوهات بالتفصيل ────────────────────────
    ws2 = wb.create_sheet("شرح السيناريوهات")
    ws2.sheet_view.rightToLeft = True

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 55

    ws2.cell(1, 1).value = "الصف"
    ws2.cell(1, 2).value = "نوع السيناريو"
    ws2.cell(1, 3).value = "التفصيل"

    for c in range(1, 4):
        ws2.cell(1, c).font = Font(bold=True, color="FFFFFF", name="Cairo")
        ws2.cell(1, c).fill = PatternFill("solid", fgColor="1E3A5F")
        ws2.cell(1, c).alignment = Alignment(horizontal="center")
        ws2.cell(1, c).border = make_border()

    scenarios = [
        (2,  "✅ نجاح",   "طالب مكتمل البيانات - اسم + هاتف + عمر + تاريخ ميلاد + صف موجود"),
        (3,  "✅ نجاح",   "طالبة مكتملة البيانات مع تاريخ ميلاد"),
        (4,  "✅ نجاح",   "هاتف مشترك مع طالب آخر (أخوه) - يُقبل لأن التكرار ليس تاماً"),
        (5,  "✅ نجاح",   "بدون هاتف ولا تاريخ ميلاد - يُقبل"),
        (6,  "✅ نجاح",   "عمر يدوي بدون تاريخ ميلاد - يُقبل"),
        (7,  "✅ نجاح",   "اسم بهمزات وتاء مربوطة - يُقبل بعد تطبيع النص"),
        (8,  "↔ يُتخطى", "صف فارغ تماماً - لا يُحسب خطأ ولا نجاحاً، يُتجاهل بصمت"),
        (9,  "❌ فشل",    "الاسم فارغ - الاسم حقل إجباري"),
        (10, "❌ فشل",    "الصف فارغ - يجب تحديد الصف"),
        (11, "❌ فشل",    "الصف غير موجود في النظام - يجب أن يكون اسم الصف مطابقاً تماماً"),
        (12, "❌ فشل",    "تكرار تام داخل الملف - نفس بيانات الصف رقم 2 بالضبط"),
        (13, "❌ فشل",    "تكرار تام مرة أخرى - يُرفض أيضاً"),
    ]

    for row_i, (row_num, stype, detail) in enumerate(scenarios, start=2):
        is_s = stype.startswith("✅")
        is_f = stype.startswith("❌")
        bg = COLOR_SUCCESS if is_s else (COLOR_FAIL if is_f else "FEF9C3")
        ws2.cell(row_i, 1).value = row_num
        ws2.cell(row_i, 2).value = stype
        ws2.cell(row_i, 3).value = detail
        for c in range(1, 4):
            ws2.cell(row_i, c).fill = PatternFill("solid", fgColor=bg)
            ws2.cell(row_i, c).border = make_border()
            ws2.cell(row_i, c).font = Font(name="Cairo", size=10)
            ws2.cell(row_i, c).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # ── حفظ الملف ────────────────────────────────────────────────
    wb.save(OUTPUT_PATH)
    print(f"\n✅ تم إنشاء الملف بنجاح:\n   {OUTPUT_PATH}\n")
    print(f"   📊 {len(data_rows) + 1} صف (الرأس + {len(data_rows)} طالب/فارغ)")
    print(f"   ✅ طلاب يجب أن ينجحوا: {sum(1 for r in ROWS if r[5].startswith('✅'))}")
    print(f"   ❌ طلاب يجب أن يفشلوا: {sum(1 for r in ROWS if r[5].startswith('❌'))}")
    print(f"\n   ⚠️  تأكد أن اسم الصف '{VALID_GRADE}' موجود في نظامك قبل الرفع!")

if __name__ == "__main__":
    build_excel()
